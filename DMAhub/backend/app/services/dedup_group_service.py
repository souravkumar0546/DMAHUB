from __future__ import annotations

import json
import os
import re
from collections import defaultdict

import pandas as pd
import httpx
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

GROUP_COLORS = [
    ("FFFF99", "FFF176"),
    ("FFCC80", "FFB74D"),
    ("A5D6A7", "81C784"),
    ("90CAF9", "64B5F6"),
    ("CE93D8", "BA68C8"),
    ("EF9A9A", "E57373"),
    ("80DEEA", "4DD0E1"),
    ("FFE082", "FFD54F"),
]


def _normalize(text: str) -> str:
    """Normalize text for duplicate detection: lowercase, strip leading/trailing
    whitespace, collapse internal whitespace, and remove non-alphanumeric chars
    (except hyphens, slashes, periods, commas which appear in catalog numbers
    and quantities)."""
    t = str(text).lower().strip()
    # Replace special chars that add noise but keep: - / . , (useful in catalog#, units)
    t = re.sub(r"[^a-z0-9\s\-/.,]", " ", t)
    # Collapse multiple spaces
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _build_tfidf_matrix(texts: list[str]):
    vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), lowercase=True)
    matrix = vec.fit_transform(texts)
    return vec, matrix


def _token_containment(a: str, b: str) -> float:
    """Fraction of tokens in the shorter string that appear in the longer one.
    Handles cases where TF-IDF underscores short-vs-long matches."""
    a, b = a.lower().strip(), b.lower().strip()
    if not a or not b:
        return 0.0
    shorter, longer = (a, b) if len(a.split()) <= len(b.split()) else (b, a)
    tokens = shorter.split()
    if not tokens:
        return 0.0

    def _token_in_longer(tok: str, longer_text: str) -> bool:
        # IMPORTANT: numeric tokens must match as whole tokens.
        # Without word boundaries, '500' is considered contained in '5000'.
        if tok.isdigit():
            return re.search(rf"\b{re.escape(tok)}\b", longer_text) is not None
        return tok in longer_text

    matched = sum(1 for t in tokens if _token_in_longer(t, longer))
    containment = matched / len(tokens)
    matched_chars = sum(len(t) for t in tokens if _token_in_longer(t, longer))
    longer_alpha = "".join(c for c in longer if c.isalnum())
    coverage = matched_chars / max(len(longer_alpha), 1)
    return containment * 0.7 + min(coverage, 1.0) * 0.3


def _pair_sim(tfidf_matrix, i: int, j: int, text_i: str, text_j: str) -> float:
    """Similarity between row i and row j for a single column.
    Uses max(TF-IDF cosine, token_containment)."""
    tfidf_score = float(cosine_similarity(tfidf_matrix[i:i + 1], tfidf_matrix[j:j + 1])[0][0])
    contain_score = _token_containment(text_i, text_j)
    return max(tfidf_score, contain_score)


def _weighted_overall(per_col_scores: dict[str, float], columns: list[str]) -> float:
    """Weighted overall: first column 60%, rest share 40%.
    None scores are excluded."""
    valid = {c: s for c, s in per_col_scores.items() if s is not None}
    if not valid:
        return 0.0
    primary = columns[0]
    if primary not in valid:
        return 0.0
    secondary = [c for c in columns[1:] if c in valid]
    if not secondary:
        return valid[primary]
    sec_avg = sum(valid[c] for c in secondary) / len(secondary)
    return valid[primary] * 0.6 + sec_avg * 0.4


def _exact_selected_match(df: pd.DataFrame, i: int, j: int, selected_cols: list[str]) -> bool:
    """True exact match across selected columns (after minimal normalization)."""
    for col in selected_cols:
        a = _normalize(df.iloc[i].get(col, ""))
        b = _normalize(df.iloc[j].get(col, ""))
        if a != b:
            return False
    return True


def _azure_ready() -> bool:
    return bool(
        os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
        and os.getenv("AZURE_OPENAI_API_KEY", "").strip()
        and os.getenv("AZURE_OPENAI_MODEL", "").strip()
    )


async def _ai_variant_check_batch_async(pairs: list[tuple[int, str, str]], timeout_s: int = 120) -> dict[int, bool]:
    """Return mapping: idx -> True if DIFFERENT SIZE/QTY (NOT duplicate)."""
    azure_ep = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/")
    azure_k = os.getenv("AZURE_OPENAI_API_KEY", "")
    azure_m = os.getenv("AZURE_OPENAI_MODEL", "")
    azure_v = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
    url = f"{azure_ep}/openai/deployments/{azure_m}/chat/completions?api-version={azure_v}"

    prompt_parts = [
        "You are a procurement/master-data deduplication expert.\n"
        "For each pair of records, decide whether they are TRUE DUPLICATES (same exact material)\n"
        "or NOT DUPLICATES because they are the same product family but DIFFERENT variant due to different\n"
        "dimension/size/quantity/pack size/strength/concentration.\n\n"
        "Rules:\n"
        "- If any size/qty/pack/strength differs (e.g., 500ml vs 1L, 10mm vs 12mm, 1000 vs 10000, 10x20 vs 10x30) -> DIFFERENT_SIZE_QTY.\n"
        "- If size/qty is missing or consistent -> DUPLICATE only if the items are otherwise the same.\n\n"
        "Return ONLY valid JSON with this shape:\n"
        '{ "0": {"verdict": "DUPLICATE" | "DIFFERENT_SIZE_QTY"}, ... }\n'
        "No other text.\n\n"
    ]
    for idx, left_text, right_text in pairs:
        prompt_parts.append(f'PAIR "{idx}"\nLEFT: {left_text}\nRIGHT: {right_text}\n\n')
    prompt = "".join(prompt_parts)

    async with httpx.AsyncClient(timeout=timeout_s) as client:
        resp = await client.post(
            url,
            headers={"Content-Type": "application/json", "api-key": azure_k},
            json={
                "messages": [
                    {"role": "system", "content": "Respond only with valid JSON."},
                    {"role": "user", "content": prompt},
                ],
                "max_tokens": 2000,
                "temperature": 0.0,
            },
        )
    resp.raise_for_status()
    text = resp.json()["choices"][0]["message"]["content"]
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```", "", text)
    parsed = json.loads(text.strip())

    out: dict[int, bool] = {}
    for idx, _l, _r in pairs:
        verdict = str(parsed.get(str(idx), {}).get("verdict", "")).strip().upper()
        out[idx] = verdict == "DIFFERENT_SIZE_QTY"
    return out


async def analyze_group_duplicates(
    df: pd.DataFrame,
    selected_cols: list[str],
    identifier_col: str,
    threshold: float,
    apply_ai_variant_filter: bool = True,
) -> list[dict]:
    n = len(df)
    col_texts: dict[str, list[str]] = {}
    for col in selected_cols:
        col_texts[col] = [_normalize(val) for val in df[col].fillna("")]

    col_tfidf: dict[str, object] = {}
    for col in selected_cols:
        try:
            _, matrix = _build_tfidf_matrix(col_texts[col])
            col_tfidf[col] = matrix
        except Exception:
            col_tfidf[col] = None

    primary = selected_cols[0]
    primary_matrix = col_tfidf.get(primary)
    if primary_matrix is None:
        return []

    pre_threshold = threshold * 0.75

    # For large datasets, process in chunks to limit memory (matches Streamlit).
    CHUNK = 500
    candidate_pairs: list[tuple[int, int]] = []

    for start in range(0, n, CHUNK):
        end = min(start + CHUNK, n)
        chunk_sim = cosine_similarity(primary_matrix[start:end], primary_matrix)

        for local_i in range(end - start):
            global_i = start + local_i
            row_sims = chunk_sim[local_i]
            for global_j in range(global_i + 1, n):
                if row_sims[global_j] >= pre_threshold:
                    candidate_pairs.append((global_i, global_j))

    pair_scores: dict[tuple[int, int], float] = {}
    for i, j in candidate_pairs:
        per_col: dict[str, float] = {}
        for col in selected_cols:
            ti = col_texts[col][i]
            tj = col_texts[col][j]
            if not ti.strip() and not tj.strip():
                per_col[col] = None
            elif not ti.strip() or not tj.strip():
                per_col[col] = 0.0
            else:
                matrix = col_tfidf.get(col)
                per_col[col] = _pair_sim(matrix, i, j, ti, tj) if matrix is not None else _token_containment(ti, tj)

        overall = _weighted_overall(per_col, selected_cols)
        if overall >= threshold:
            pair_scores[(i, j)] = overall

    neighbors: dict[int, list[tuple[int, float]]] = defaultdict(list)
    for (i, j), s in pair_scores.items():
        neighbors[i].append((j, s))
        neighbors[j].append((i, s))

    out_rows: list[dict] = []
    shown_pairs: set[tuple[int, int]] = set()
    ai_checks: list[tuple[int, str, str]] = []
    block_id = 0

    for i in range(n):
        if i not in neighbors:
            continue
        matches = sorted(neighbors[i], key=lambda x: (-x[1], x[0]))
        if not matches:
            continue

        has_new_pair = False
        for j, _score in matches:
            a, b = (i, j) if i < j else (j, i)
            if (a, b) not in shown_pairs:
                has_new_pair = True
                break
        if not has_new_pair:
            continue

        for k, (j, score) in enumerate(matches):
            a, b = (i, j) if i < j else (j, i)
            shown_pairs.add((a, b))
            row_out: dict = {}
            row_out["block_id"] = block_id

            row_out[f"Left {identifier_col}"] = df.iloc[i].get(identifier_col, "") if k == 0 else ""
            for c in selected_cols:
                row_out[f"Left {c}"] = df.iloc[i].get(c, "") if k == 0 else ""

            row_out[f"Right {identifier_col}"] = df.iloc[j].get(identifier_col, "")
            for c in selected_cols:
                row_out[f"Right {c}"] = df.iloc[j].get(c, "")

            row_out["Match %"] = round(float(score * 100.0), 2)
            if _exact_selected_match(df, i, j, selected_cols):
                row_out["Verdict"] = "EXACT MATCH"
            elif score >= 0.90:
                row_out["Verdict"] = "VERY LIKELY DUPLICATE"
            else:
                row_out["Verdict"] = "POTENTIAL DUPLICATE"

            out_rows.append(row_out)
            if row_out["Verdict"] != "EXACT MATCH":
                left_text = " | ".join(f"{c}: {df.iloc[i].get(c, '')}" for c in selected_cols)
                right_text = " | ".join(f"{c}: {df.iloc[j].get(c, '')}" for c in selected_cols)
                ai_checks.append((len(out_rows) - 1, left_text, right_text))

        block_id += 1

    if apply_ai_variant_filter and ai_checks and _azure_ready():
        batch_size = 20
        for start in range(0, len(ai_checks), batch_size):
            batch = ai_checks[start:start + batch_size]
            indexed = [(idx, l, r) for idx, (_out_i, l, r) in enumerate(batch)]
            try:
                verdicts = await _ai_variant_check_batch_async(indexed)
                for local_idx, (out_i, _l, _r) in enumerate(batch):
                    if verdicts.get(local_idx, False):
                        out_rows[out_i]["Verdict"] = "NOT DUPLICATE (DIFFERENT SIZE/QTY)"
            except Exception:
                pass

    out_rows = [r for r in out_rows if r.get("Verdict") != "NOT DUPLICATE (DIFFERENT SIZE/QTY)"]
    for r in out_rows:
        for k in list(r.keys()):
            if k.startswith("Left ") or k.startswith("Right ") or k == "Verdict":
                r[k] = str(r[k])
            elif k == "block_id":
                r[k] = int(r[k])
    return out_rows


def build_duplicate_report_colored_xlsx(rows: list[dict]) -> bytes:
    """Colored XLSX: alternating fills per block_id."""
    if not rows:
        raise ValueError("rows is empty")
    export_rows = []
    for r in rows:
        er = {k: v for k, v in r.items() if k != "block_id"}
        export_rows.append(er)
    export_df = pd.DataFrame(export_rows)
    block_ids = [int(r.get("block_id", i)) for i, r in enumerate(rows)]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Duplicate Report")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb["Duplicate Report"]
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(bottom=Side(style="thin", color="999999"))
    col_count = len(export_df.columns)

    for ci in range(1, col_count + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for out_i, excel_ri in enumerate(range(2, len(export_df) + 2)):
        bid = block_ids[out_i] if out_i < len(block_ids) else out_i
        pair = GROUP_COLORS[bid % len(GROUP_COLORS)]
        fill_color = pair[0] if (bid % 2 == 0) else pair[1]
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for ci in range(1, col_count + 1):
            cell = ws.cell(row=excel_ri, column=ci)
            cell.fill = fill
            cell.border = thin_border

    for ci in range(1, col_count + 1):
        col_letter = ws.cell(row=1, column=ci).column_letter
        max_len = len(str(ws.cell(row=1, column=ci).value or ""))
        for ri in range(2, min(len(export_df) + 2, 400)):
            val = ws.cell(row=ri, column=ci).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
