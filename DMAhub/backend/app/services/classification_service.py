from __future__ import annotations

import json
import os
import re
from collections import Counter, defaultdict
from io import BytesIO
from typing import Optional

import numpy as np
import pandas as pd
import httpx
from nltk.stem import SnowballStemmer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

stemmer = SnowballStemmer("english")

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

BRANDS = {
    "sigma", "aldrich", "merck", "borosil", "rankem", "tarson", "tarsons",
    "corning", "invitrogen", "thermo", "fisher", "scientific", "biorad",
    "bio-rad", "waters", "agilent", "millipore", "pall", "sarstedt",
    "eppendorf", "bd", "becton", "dickinson", "hamilton", "make",
    "brand", "honeywell", "labserv", "vwr", "duran", "nalgene", "abcam",
    "biolegend", "gentest", "perkinelmer",
}

import copy as _copy
import pathlib as _pathlib
from app.services.taxonomy_data import MATERIAL_TAXONOMIES, TAXONOMY_OPTIONS

# Default taxonomy key
DEFAULT_TAXONOMY_KEY = "ZSC1"

# ── Disk persistence for user-edited taxonomies ──
_PERSIST_DIR = _pathlib.Path(__file__).resolve().parent / "_taxonomy_overrides"
_PERSIST_DIR.mkdir(exist_ok=True)


def _persist_path(key: str) -> _pathlib.Path:
    return _PERSIST_DIR / f"{key}.json"


def _load_taxonomy_for_key(key: str) -> dict:
    """Load taxonomy: from disk override if exists, otherwise from hardcoded defaults."""
    path = _persist_path(key)
    if path.exists():
        try:
            saved = json.loads(path.read_text(encoding="utf-8"))
            if "categories" in saved:
                return saved
        except Exception:
            pass
    # Fall back to hardcoded
    if key not in MATERIAL_TAXONOMIES:
        key = DEFAULT_TAXONOMY_KEY
    tax = MATERIAL_TAXONOMIES[key]
    return {
        "taxonomy_key": key,
        "categories": _copy.deepcopy(tax["categories"]),
        "override_rules": tax["override_rules"],
    }


def _save_taxonomy_to_disk(key: str, data: dict):
    """Persist user edits so they survive backend restarts."""
    path = _persist_path(key)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Mutable taxonomy state (loaded from disk or defaults) ──
_taxonomy_store: dict = _load_taxonomy_for_key(DEFAULT_TAXONOMY_KEY)
_taxonomy_store.setdefault("taxonomy_key", DEFAULT_TAXONOMY_KEY)


def get_taxonomy() -> dict:
    """Return current taxonomy config."""
    return _copy.deepcopy(_taxonomy_store)


def get_all_taxonomies() -> list[dict]:
    """Return the list of available taxonomy options for the dropdown."""
    return TAXONOMY_OPTIONS


def set_taxonomy(data: dict) -> dict:
    """Update taxonomy config. Accepts taxonomy_key to switch, or categories/override_rules to edit."""
    switching_key = False
    # If a taxonomy_key is provided, load that taxonomy (from disk or defaults)
    if "taxonomy_key" in data and data["taxonomy_key"] in MATERIAL_TAXONOMIES:
        key = data["taxonomy_key"]
        loaded = _load_taxonomy_for_key(key)
        _taxonomy_store["taxonomy_key"] = key
        _taxonomy_store["categories"] = loaded["categories"]
        _taxonomy_store["override_rules"] = loaded["override_rules"]
        switching_key = True
    # Allow per-field overrides on top (user editing definitions/rules)
    edited = False
    if "categories" in data and isinstance(data["categories"], dict) and not switching_key:
        _taxonomy_store["categories"] = data["categories"]
        edited = True
    if "override_rules" in data and isinstance(data["override_rules"], str) and not switching_key:
        _taxonomy_store["override_rules"] = data["override_rules"]
        edited = True
    # Persist edits to disk so they survive restarts
    if edited:
        key = _taxonomy_store.get("taxonomy_key", DEFAULT_TAXONOMY_KEY)
        _save_taxonomy_to_disk(key, {
            "taxonomy_key": key,
            "categories": _taxonomy_store["categories"],
            "override_rules": _taxonomy_store["override_rules"],
        })
        # Also update the in-memory MATERIAL_TAXONOMIES so switching back loads the edits
        if key in MATERIAL_TAXONOMIES:
            MATERIAL_TAXONOMIES[key]["categories"] = _copy.deepcopy(_taxonomy_store["categories"])
            MATERIAL_TAXONOMIES[key]["override_rules"] = _taxonomy_store["override_rules"]
    # Keep module-level list in sync
    global CONSUMABLE_L2_CATEGORIES
    CONSUMABLE_L2_CATEGORIES = list(_taxonomy_store["categories"].keys())
    return _copy.deepcopy(_taxonomy_store)


def _get_category_defs() -> dict:
    return _taxonomy_store["categories"]


def get_taxonomy_for_key(key: str) -> dict:
    """Load taxonomy for a specific key without changing the global store."""
    return _load_taxonomy_for_key(key)


CONSUMABLE_L2_CATEGORIES = list(_taxonomy_store["categories"].keys())


# ══════════════════════════════════════════════════════════════════════════════
#  AZURE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _azure_cfg():
    ep = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/")
    key = os.getenv("AZURE_OPENAI_API_KEY", "")
    model = os.getenv("AZURE_OPENAI_MODEL", "")
    ver = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
    return ep, key, model, ver


async def _azure_chat_json(user_prompt: str, system_prompt: str, max_tokens: int = 4096, temperature: float = 0.1, timeout: int = 120):
    ep, key, model, ver = _azure_cfg()
    url = f"{ep}/openai/deployments/{model}/chat/completions?api-version={ver}"
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            url,
            headers={"Content-Type": "application/json", "api-key": key},
            json={
                "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                "max_tokens": max_tokens,
                "temperature": temperature,
            },
            timeout=timeout,
        )
        resp.raise_for_status()
        text = resp.json()["choices"][0]["message"]["content"]
        text = re.sub(r"```json\s*", "", text)
        text = re.sub(r"```", "", text).strip()
        return json.loads(text)


# ══════════════════════════════════════════════════════════════════════════════
#  TAXONOMY RUBRIC & CATEGORY MATCHING
# ══════════════════════════════════════════════════════════════════════════════

def _build_taxonomy_rubric() -> str:
    cat_defs = _get_category_defs()
    overrides = _taxonomy_store["override_rules"]
    lines = [
        "CLASSIFICATION TAXONOMY — L2 Categories under Consumables\n"
        "==========================================================\n"
        "Use these definitions and examples to classify materials. "
        "Match based on the material's FUNCTION and USE, not just its name.\n\n"
    ]
    for i, (cat, info) in enumerate(cat_defs.items(), 1):
        lines.append(f"{i}. **{cat}**\n")
        lines.append(f"   Definition: {info['definition']}\n")
        lines.append(f"   Examples: {info['examples']}\n")
        lines.append("\n")
    lines.append(
        "DISAMBIGUATION RULES:\n"
        "- If a material is used in analytical TESTING → Analytical Consumables\n"
        "- If a material is consumed in MANUFACTURING/PROCESS → Process Consumables\n"
        "- If it is PPE or SAFETY gear → Safety Consumables\n"
        "- If it is for EQUIPMENT REPAIR/MAINTENANCE → Engineering & Maintenance Consumables\n"
        "- If it is general-purpose LAB use (glassware, tips, storage) → Laboratory Supplies\n"
        "- If it is a CULTURE MEDIA or BUFFER solution → Media & Buffers\n"
        "- If it is a chemical WITHOUT a CAS number (catalogue-only) → Non-CAS based Chemicals\n"
        "- If it is a REFERENCE STANDARD or RLD → RLD & Reference Standards\n"
        "- If it is for FACILITY/PANTRY/CLEANING (non-lab) → House Keeping Supplies\n"
        "- If it is a MEDICINE or DRUG product → Medicinal Supplies\n"
        "- If it is a QC REAGENT, CALIBRATOR, or ANTIBODY → Lab Reagents & QC Supplies\n"
        "- If it involves DNA, RNA, PROTEIN, GENE work → Molecular Biology Consumables\n"
        "\n"
        f"CRITICAL OVERRIDE RULES (apply these FIRST, before the rules above):\n{overrides}\n"
    )
    return "".join(lines)


def _match_category_flexible(cat: str) -> Optional[str]:
    ALIASES = {
        "lab reagents and quality check": "Lab Reagents & QC Supplies",
        "lab reagents & quality check": "Lab Reagents & QC Supplies",
        "molecular biology": "Molecular Biology Consumables",
        "house keeping supplies": "House Keeping Supplies",
        "housekeeping supplies": "House Keeping Supplies",
        "mole biology cons": "Molecular Biology Consumables",
        "mole biology consumables": "Molecular Biology Consumables",
    }
    allowed_cats = list(_get_category_defs().keys())
    if cat in allowed_cats:
        return cat
    cat_lower = cat.strip().lower()
    if cat_lower in ALIASES:
        return ALIASES[cat_lower]
    for allowed in allowed_cats:
        if allowed.lower() == cat_lower:
            return allowed
    for allowed in allowed_cats:
        if allowed.lower() in cat_lower or cat_lower in allowed.lower():
            return allowed
    return None


def _find_key_in_parsed(parsed: dict, grp_num: int) -> Optional[dict]:
    candidates = [
        str(grp_num), f"GROUP {grp_num}", f"Group {grp_num}", f"group {grp_num}",
        f"G{grp_num}", f"group_{grp_num}", f"GROUP_{grp_num}", f"Group_{grp_num}",
    ]
    for key in candidates:
        if key in parsed:
            return parsed[key]
    for key, val in parsed.items():
        m = re.search(r"(\d+)\s*$", str(key))
        if m and int(m.group(1)) == grp_num:
            return val
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  TEXT CLEANING & STEMMING
# ══════════════════════════════════════════════════════════════════════════════

def clean_description(text: str) -> str:
    text = str(text).lower().strip()
    # Strip catalog numbers / part numbers / SKUs
    text = re.sub(
        r"(cat\.?\s*#?\s*:?\s*no\.?\s*:?\s*|cat\s*#\s*:?\s*|p/?n\s*:?\s*#?\s*-?\s*"
        r"|code\s*:?\s*-?\s*|pn\s*#\s*:?\s*|sku\s*:?\s*|ref\s*:?\s*|art\.?\s*:?\s*"
        r"|part\s*no\.?\s*:?\s*-?\s*)\s*[\w\-\.]+",
        " ", text,
    )
    # Strip standalone catalog-like codes (4+ digit alphanumeric)
    text = re.sub(r"\b[a-z]?\d{4,}[\w\-]*\b", " ", text)
    # Strip specific number+unit patterns
    text = re.sub(r"\b\d+\s*mil\b", " ", text)
    text = re.sub(r"\b\d+g\b", " ", text)
    text = re.sub(r"\bb-?\d+/?\.?\d*\b", " ", text)
    # Strip quantity + unit (expanded unit list matching Streamlit)
    text = re.sub(
        r"\b\d+[\.,]?\d*\s*"
        r"(ml|mg|l|g|kg|mm|cm|m|µl|ul|oz|lb|pcs|pc|ea|x|inch|in|ft|cc"
        r"|µm|nm|um|mcg|iu|units?|ltr|litre|liter|gal|gallon|qt|pt"
        r"|mm2|cm2|m2|kgs?|ltrs?|µ)\b",
        " ", text,
    )
    # Strip remaining standalone numbers
    text = re.sub(r"\b\d+[\.,]?\d*\b", " ", text)
    # Strip brand names
    for brand in BRANDS:
        text = re.sub(r"\b" + re.escape(brand) + r"\b", " ", text)
    text = re.sub(r"[^a-z\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def stem_and_sort(text: str) -> str:
    tokens = [t for t in text.split() if len(t) > 1]
    return " ".join(sorted(set(stemmer.stem(t) for t in tokens)))


def generate_blocking_keys(cleaned: str, stemmed: str) -> list[str]:
    keys = set()
    if stemmed.strip():
        keys.add(stemmed)
    STOPWORDS = {"with", "for", "and", "the", "from", "type", "set", "cat", "no", "dia", "size"}
    tokens = cleaned.split()
    significant = [t for t in tokens if t not in STOPWORDS and len(t) > 2]
    if len(significant) >= 2:
        sig_stemmed = sorted(set(stemmer.stem(t) for t in significant[:3]))
        keys.add(" ".join(sig_stemmed))
    if significant:
        core = max(significant, key=len)
        core_stem = stemmer.stem(core)
        if len(core_stem) >= 5:
            keys.add(core_stem)
    return [k for k in keys if k.strip()]


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 1 — BUILD GROUPS (Union-Find)
# ══════════════════════════════════════════════════════════════════════════════

def build_semantic_groups(descriptions: list[str], threshold: float) -> dict[int, list[int]]:
    n = len(descriptions)
    cleaned = [clean_description(d) for d in descriptions]
    stemmed = [stem_and_sort(c) for c in cleaned]

    candidate_blocks = defaultdict(set)
    for i in range(n):
        for k in generate_blocking_keys(cleaned[i], stemmed[i]):
            candidate_blocks[k].add(i)

    MULTI_WORD_CAP = 100
    SINGLE_WORD_CAP = 20
    candidate_groups = []
    for key, members in candidate_blocks.items():
        if len(members) < 2:
            continue
        is_single = " " not in key
        cap = SINGLE_WORD_CAP if is_single else MULTI_WORD_CAP
        if len(members) > cap:
            continue
        candidate_groups.append(list(members))

    parent = list(range(n))
    rank_arr = [0] * n

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        if rank_arr[ra] < rank_arr[rb]:
            ra, rb = rb, ra
        parent[rb] = ra
        if rank_arr[ra] == rank_arr[rb]:
            rank_arr[ra] += 1

    for members in candidate_groups:
        sub_texts = [cleaned[m] for m in members]
        if len(set(sub_texts)) == 1:
            for m in members[1:]:
                union(members[0], m)
            continue
        try:
            tfidf = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), lowercase=True).fit_transform(sub_texts)
            sim = cosine_similarity(tfidf)
            for i in range(len(members)):
                for j in range(i + 1, len(members)):
                    if sim[i, j] >= threshold:
                        union(members[i], members[j])
        except Exception:
            for m in members[1:]:
                union(members[0], m)

    groups = defaultdict(list)
    for i in range(n):
        groups[find(i)].append(i)
    return {k: v for k, v in groups.items() if len(v) >= 2}


# ══════════════════════════════════════════════════════════════════════════════
#  AI VALIDATION (outlier removal)
# ══════════════════════════════════════════════════════════════════════════════

async def validate_groups_batch_with_ai(batch: list, columns: list[str]) -> dict:
    prompt_parts = [
        "You are a pharma / lab materials classification expert.\n"
        "Below are MULTIPLE groups of material records. Each group was clustered "
        "by text similarity. Some items may be outliers that don't belong.\n\n"
        "Review ALL columns for each group and decide which items genuinely "
        "belong together in the same material family.\n\n"
    ]
    for gkey, rows in batch:
        prompt_parts.append(f'--- GROUP "{gkey}" ---\n')
        for idx, row in enumerate(rows):
            fields = " | ".join(f"{col}: {row.get(col, '')}" for col in columns)
            prompt_parts.append(f"  {idx}. {fields}\n")
        prompt_parts.append("\n")
    prompt_parts.append(
        "Return ONLY a JSON object where each key is the group name and "
        "each value is an object with:\n"
        '  "valid": [indices that belong together]\n'
        '  "outliers": [indices that do NOT belong]\n'
        'Example: {"GROUP_A": {"valid": [0,1,2], "outliers": [3]}, ...}\n'
        "No other text."
    )
    prompt = "".join(prompt_parts)
    defaults = {gkey: list(range(len(rows))) for gkey, rows in batch}
    try:
        parsed = await _azure_chat_json(
            prompt,
            "You are a pharma materials classification expert. Respond only with valid JSON.",
        )
        result = {}
        for gkey, rows in batch:
            if gkey in parsed and "valid" in parsed[gkey]:
                result[gkey] = parsed[gkey]["valid"]
            else:
                result[gkey] = defaults[gkey]
        return result
    except Exception:
        return defaults


# ══════════════════════════════════════════════════════════════════════════════
#  AI RECLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════

def _majority_vote(classifications: list) -> str:
    if not classifications:
        return ""
    return Counter(classifications).most_common(1)[0][0]


async def reclassify_groups_batch_with_ai(batch: list, desc_col: str, po_col: Optional[str]) -> dict:
    prompt_parts = [
        "You are a pharma/lab materials classification expert.\n\n"
        f"{_build_taxonomy_rubric()}\n"
        "TASK: For each group below, I give you material descriptions.\n"
        "Step 1: Identify WHAT this material is (the noun/identity, e.g. 'pipette tips', 'HPLC column', 'nitrile gloves').\n"
        "Step 2: Based on that identity, classify into EXACTLY one of the 12 categories above.\n"
        "You MUST choose exactly one category from this list and return its name EXACTLY as written in the rubric. Never invent a new category name.\n\n"
    ]
    for grp_num, rows, _ in batch:
        # Deduplicate rows within each group
        seen = set()
        entries = []
        for row in rows:
            desc = str(row.get(desc_col, "")).strip()
            po = str(row.get(po_col, "")).strip() if po_col else ""
            key = (desc, po)
            if key not in seen and desc:
                seen.add(key)
                entries.append((desc, po))
        prompt_parts.append(f"--- GROUP {grp_num} ---\n")
        for i, (desc, po) in enumerate(entries):
            if po and po != desc and po != "nan":
                prompt_parts.append(f"  {i}. {desc} | PO: {po}\n")
            else:
                prompt_parts.append(f"  {i}. {desc}\n")
        prompt_parts.append("\n")
    prompt_parts.append(
        "Return ONLY a JSON object. Each key is the group number as a plain string.\n"
        "IMPORTANT: Return ONE classification per GROUP, NOT per row. All rows in a group get the SAME category.\n"
        "Each value must have:\n"
        '  "material_identity": "<what this material IS, 2-5 words>",\n'
        '  "category": "<EXACTLY one of the 12 categories above>",\n'
        '  "confidence": <integer 0-100>,\n'
        '  "rationale": "<brief explanation, max 150 chars>"\n\n'
        'Example: {"1": {"material_identity": "pipette tips", "category": "Laboratory Supplies", "confidence": 95, "rationale": "Pipette tips are general lab consumables"}}\n'
        "The category field MUST be one of the 12 categories from the rubric above; do not output any new or modified category names.\n"
        "No other text outside the JSON.\n"
    )
    prompt = "".join(prompt_parts)

    parsed = None
    try:
        ep, key, model, ver = _azure_cfg()
        url = f"{ep}/openai/deployments/{model}/chat/completions?api-version={ver}"
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                url,
                headers={"Content-Type": "application/json", "api-key": key},
                json={
                    "messages": [
                        {"role": "system", "content": "You are a pharma materials classification expert. Respond ONLY with valid JSON. First identify what the material IS, then classify it. Use group number as plain string key."},
                        {"role": "user", "content": prompt},
                    ],
                    "max_tokens": 4096,
                    "temperature": 0.1,
                },
                timeout=120,
            )
            if resp.status_code == 200:
                raw_text = resp.json()["choices"][0]["message"]["content"]
                cleaned_text = re.sub(r"```json\s*", "", raw_text)
                cleaned_text = re.sub(r"```", "", cleaned_text).strip()
                parsed = json.loads(cleaned_text)
    except Exception:
        pass

    result = {}
    for grp_num, rows, current_cats in batch:
        matched = None
        if parsed is not None:
            matched = _find_key_in_parsed(parsed, grp_num)

        if matched and isinstance(matched, dict):
            # Handle case where AI returned per-row instead of per-group
            if "category" not in matched and all(isinstance(v, dict) for v in matched.values()):
                row_cats = [v.get("category", "") for v in matched.values() if isinstance(v, dict) and v.get("category")]
                if row_cats:
                    most_common = Counter(row_cats).most_common(1)[0][0]
                    sample_identity = next((v.get("material_identity", "") for v in matched.values() if isinstance(v, dict)), "")
                    sample_rationale = next((v.get("rationale", "") for v in matched.values() if isinstance(v, dict)), "")
                    matched = {"category": most_common, "confidence": 80, "material_identity": sample_identity, "rationale": f"(per-row vote) {sample_rationale}"}

            if "category" in matched:
                raw_cat = str(matched["category"]).strip()
                conf = matched.get("confidence", 0)
                identity = str(matched.get("material_identity", ""))[:100]
                rationale = str(matched.get("rationale", ""))[:200]
                resolved_cat = _match_category_flexible(raw_cat)
                if resolved_cat:
                    result[grp_num] = {"category": resolved_cat, "confidence": conf, "rationale": f"[{identity}] {rationale}"}
                    continue

        majority = _majority_vote(current_cats)
        result[grp_num] = {"category": majority, "confidence": 50, "rationale": f"Majority vote fallback: {majority}"}

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 1 — FIX MASTER DATA
# ══════════════════════════════════════════════════════════════════════════════

async def fix_master_data(df_master: pd.DataFrame, desc_col: str, class_col: str, po_col: Optional[str], status_col: Optional[str], sim_threshold: float, ai_batch_size: int) -> tuple[pd.DataFrame, dict]:
    original_snapshot = df_master.copy()
    working = df_master.copy()
    if status_col:
        active = working[status_col].astype(str).str.strip().str.lower() == "active"
    else:
        active = pd.Series(True, index=working.index)
    mask = active & working[desc_col].notna()
    df_valid = working[mask].reset_index(drop=True)
    original_indices = working[mask].index.tolist()
    descriptions = df_valid[desc_col].astype(str).tolist()
    classifications = df_valid[class_col].astype(str).tolist()
    groups = build_semantic_groups(descriptions, sim_threshold)
    groups = {k: v for k, v in groups.items() if len(v) >= 2}
    inconsistent = [gk for gk, m in groups.items() if len(set(classifications[i] for i in m)) > 1]

    all_columns = [c for c in df_valid.columns]
    validated_map = {}
    for bi in range(0, len(inconsistent), ai_batch_size):
        batch_keys = inconsistent[bi:bi + ai_batch_size]
        batch = []
        for gk in batch_keys:
            rows = [{col: str(df_valid.iloc[m].get(col, "")) for col in all_columns} for m in groups[gk]]
            batch.append((f"G{gk}", rows))
        validated_map.update(await validate_groups_batch_with_ai(batch, all_columns))

    for gk in list(inconsistent):
        members = groups[gk]
        valid_local = validated_map.get(f"G{gk}", list(range(len(members))))
        refined = [members[v] for v in valid_local if v < len(members)]
        if len(refined) >= 2:
            groups[gk] = refined
        else:
            del groups[gk]

    ordered = []
    num = 0
    for gk in sorted(groups.keys(), key=lambda x: -len(groups[x])):
        members = groups[gk]
        if len(set(classifications[m] for m in members)) > 1:
            num += 1
            ordered.append((num, members))

    reclass_map = {}
    for bi in range(0, len(ordered), ai_batch_size):
        chunk = ordered[bi:bi + ai_batch_size]
        batch = []
        for grp_num, members in chunk:
            rows = [{col: str(df_valid.iloc[m].get(col, "")) for col in all_columns} for m in members]
            cats = [classifications[m] for m in members]
            batch.append((grp_num, rows, cats))
        result = await reclassify_groups_batch_with_ai(batch, desc_col, po_col)
        grp_to_members = {gn: mems for gn, mems in chunk}
        for grp_num, rc in result.items():
            for m in grp_to_members.get(grp_num, []):
                orig_idx = original_indices[m]
                reclass_map[orig_idx] = rc

    # Build the "New Classification" column — blank by default, only filled for changed rows
    new_col = "New Classification"
    working[new_col] = ""

    corrected = 0
    for orig_idx, rc in reclass_map.items():
        new_cat = rc["category"]
        if str(working.at[orig_idx, class_col]) != str(new_cat):
            working.at[orig_idx, new_col] = new_cat
            corrected += 1

    # Reorder so "New Classification" appears right after the original class_col
    cols = list(working.columns)
    cols.remove(new_col)
    insert_pos = cols.index(class_col) + 1
    cols.insert(insert_pos, new_col)
    working = working[cols]

    changed_indices = [
        i for i in range(len(working))
        if str(working.iloc[i][new_col]).strip() != ""
    ]

    summary = {
        "groups": len(groups),
        "inconsistent_groups": len(ordered),
        "rows_updated": corrected,
        "changed_row_indices": changed_indices,
        "new_col": new_col,
    }
    return working, summary


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 2 — CLASSIFY NEW MATERIAL
# ══════════════════════════════════════════════════════════════════════════════

def _build_index_artifacts(df_master: pd.DataFrame, desc_col: str, class_col: str, sim_threshold: float):
    descriptions = df_master[desc_col].astype(str).tolist()
    # Use "New Classification" if non-empty (corrected by Phase 1), else fall back to original
    new_col = "New Classification"
    if new_col in df_master.columns:
        classifications = [
            str(row[new_col]).strip() if str(row[new_col]).strip() else str(row[class_col])
            for _, row in df_master.iterrows()
        ]
    else:
        classifications = df_master[class_col].astype(str).tolist()
    groups = build_semantic_groups(descriptions, sim_threshold)
    cleaned = [clean_description(d) for d in descriptions]
    stemmed = [stem_and_sort(c) for c in cleaned]
    combined = [f"{c} {s}" for c, s in zip(cleaned, stemmed)]
    vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), lowercase=True)
    mat = vec.fit_transform(combined)
    cat_examples: dict[str, list[str]] = {}
    for d, c in zip(descriptions, classifications):
        cat_examples.setdefault(c, [])
        if len(cat_examples[c]) < 5:
            cat_examples[c].append(d[:100])
    return groups, cleaned, stemmed, descriptions, classifications, vec, mat, cat_examples


def find_matching_material(new_text, groups, cleaned_master, stemmed_master, descriptions, classifications, tfidf_vectorizer, tfidf_matrix, threshold):
    new_cleaned = clean_description(new_text)
    new_stemmed = stem_and_sort(new_cleaned)
    new_keys = set(generate_blocking_keys(new_cleaned, new_stemmed))
    candidate_group_keys = set()
    for gk, members in groups.items():
        for m in members:
            if set(generate_blocking_keys(cleaned_master[m], stemmed_master[m])) & new_keys:
                candidate_group_keys.add(gk)
                break
    best_group = None
    best_score = 0.0
    best_matches = []
    for gk in candidate_group_keys:
        members = groups[gk]
        sub = [cleaned_master[m] for m in members] + [new_cleaned]
        try:
            tfidf = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), lowercase=True).fit_transform(sub)
            sims = cosine_similarity(tfidf[-1:], tfidf[:-1]).flatten()
            mx = float(np.max(sims))
            if mx >= threshold and mx > best_score:
                best_score = mx
                best_group = gk
                best_matches = []
                for idx in np.argsort(sims)[::-1][:5]:
                    if sims[idx] >= threshold:
                        m = members[idx]
                        best_matches.append({"description": descriptions[m], "category": classifications[m], "score": float(sims[idx])})
        except Exception:
            pass
    if best_group is not None:
        return best_matches, classifications[groups[best_group][0]]
    sims = cosine_similarity(tfidf_vectorizer.transform([f"{new_cleaned} {new_stemmed}"]), tfidf_matrix).flatten()
    fallback = []
    for idx in np.argsort(sims)[::-1][:5]:
        if float(sims[idx]) >= threshold:
            fallback.append({"description": descriptions[idx], "category": classifications[idx], "score": float(sims[idx])})
    if fallback:
        return fallback, fallback[0]["category"]
    return [], None


async def ai_confirm_match(new_desc, new_po, matched_rows, group_category):
    ep, key, model, ver = _azure_cfg()
    url = f"{ep}/openai/deployments/{model}/chat/completions?api-version={ver}"
    matched_list = "\n".join(
        f"  - {r['description']} (category: {r['category']}, similarity: {r['score']:.2f})"
        for r in matched_rows[:5]
    )
    prompt = (
        "You are a pharma / lab materials classification expert.\n\n"
        "A new material needs to be classified:\n"
        f"- Material Description: {new_desc}\n"
        f"- PO Text: {new_po}\n\n"
        "The system found these similar materials in the master data, "
        f"all classified as **{group_category}**:\n{matched_list}\n\n"
        "Question: Is the new material truly the SAME type of material as the "
        "matched ones? Should it get the same category?\n\n"
        "Return ONLY a JSON object:\n"
        '{"match": true/false, "reasoning": "<brief explanation>"}\n'
        "No other text."
    )
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                url,
                headers={"Content-Type": "application/json", "api-key": key},
                json={
                    "messages": [
                        {"role": "system", "content": "You are a pharma materials classification expert. Respond only with valid JSON."},
                        {"role": "user", "content": prompt},
                    ],
                    "max_tokens": 256,
                    "temperature": 0,
                },
                timeout=60,
            )
            resp.raise_for_status()
            text = resp.json()["choices"][0]["message"]["content"]
            text = re.sub(r"```json\s*", "", text)
            text = re.sub(r"```", "", text)
            return json.loads(text.strip())
    except Exception as e:
        return {"match": True, "reasoning": f"AI call failed ({e}), defaulting to match."}


async def ai_classify_new(new_desc, new_po):
    ep, key, model, ver = _azure_cfg()
    url = f"{ep}/openai/deployments/{model}/chat/completions?api-version={ver}"
    rubric = _build_taxonomy_rubric()
    allowed = ", ".join(CONSUMABLE_L2_CATEGORIES)
    prompt = (
        "You are a pharma/lab materials classification expert.\n\n"
        f"{rubric}\n\n"
        "TASK: Identify what this material IS, then classify into EXACTLY one of the L2 categories above.\n"
        f"ALLOWED CATEGORIES (use exact name): {allowed}\n"
        "You MUST choose exactly one category from this list and return its name EXACTLY as written. Never invent a new category name.\n\n"
        f"Material Description: {new_desc}\n"
        f"PO Text: {new_po}\n\n"
        "Return ONLY a JSON object with:\n"
        '  "material_identity": "<what this material IS, 2-5 words>",\n'
        '  "category": "<EXACTLY one of the allowed categories above>",\n'
        '  "confidence": <integer 0-100>,\n'
        '  "rationale": "<brief explanation, max 150 chars>"\n'
        "The category field MUST be one of the allowed categories; do not output any new or modified category names.\n"
        "No other text."
    )
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                url,
                headers={"Content-Type": "application/json", "api-key": key},
                json={
                    "messages": [
                        {"role": "system", "content": "You are a pharma materials classification expert. Respond ONLY with valid JSON. Use the rubric to choose ONE L2 category."},
                        {"role": "user", "content": prompt},
                    ],
                    "max_tokens": 512,
                    "temperature": 0.1,
                },
                timeout=60,
            )
            resp.raise_for_status()
            text = resp.json()["choices"][0]["message"]["content"]
            text = re.sub(r"```json\s*", "", text)
            text = re.sub(r"```", "", text)
            parsed = json.loads(text.strip())
            raw_cat = str(parsed.get("category", "")).strip()
            resolved = _match_category_flexible(raw_cat) or "UNKNOWN"
            return {"category": resolved, "confidence": parsed.get("confidence", 0), "reasoning": parsed.get("rationale", "")}
    except Exception as e:
        return {"category": "UNKNOWN", "confidence": 0, "reasoning": f"AI failed: {e}"}


async def classify_new_material(df_master: pd.DataFrame, desc_col: str, class_col: str, input_values: dict[str, str], sim_threshold: float):
    search_text = " ".join(v for v in input_values.values() if str(v).strip())
    if not search_text.strip():
        return {"error": "Please fill in at least one field."}
    groups, cleaned, stemmed, descriptions, classifications, vec, mat, _cat_examples = _build_index_artifacts(df_master, desc_col, class_col, sim_threshold)
    matched_rows, group_cat = find_matching_material(search_text, groups, cleaned, stemmed, descriptions, classifications, vec, mat, sim_threshold)
    desc_value = input_values.get(desc_col, search_text)
    po_value = " ".join(v for k, v in input_values.items() if k != desc_col and str(v).strip())
    if matched_rows and group_cat:
        confirm = await ai_confirm_match(desc_value, po_value, matched_rows, group_cat)
        if confirm.get("match", False):
            return {
                "assigned_category": group_cat,
                "method": "group_match_confirmed",
                "reasoning": confirm.get("reasoning", ""),
                "matched_rows": matched_rows,
            }
    ai_result = await ai_classify_new(desc_value, po_value)
    ac = ai_result.get("category", "UNKNOWN")
    return {
        "assigned_category": ac,
        "method": "ai_new_classification",
        "reasoning": ai_result.get("reasoning", ""),
        "matched_rows": matched_rows,
    }


async def classify_batch(
    df_master: pd.DataFrame,
    desc_col: str,
    class_col: str,
    items: list[dict[str, str]],
    sim_threshold: float,
) -> list[dict]:
    """Classify multiple items against master. Builds TF-IDF index once for performance."""
    groups, cleaned, stemmed, descriptions, classifications, vec, mat, _cat_examples = \
        _build_index_artifacts(df_master, desc_col, class_col, sim_threshold)

    results = []
    for item in items:
        search_text = " ".join(v for v in item.values() if str(v).strip())
        if not search_text.strip():
            results.append({"assigned_category": "UNKNOWN", "method": "empty_input", "reasoning": "Empty input", "matched_rows": []})
            continue
        matched_rows, group_cat = find_matching_material(
            search_text, groups, cleaned, stemmed, descriptions, classifications, vec, mat, sim_threshold,
        )
        desc_value = item.get(desc_col, search_text)
        po_value = " ".join(v for k, v in item.items() if k != desc_col and str(v).strip())
        if matched_rows and group_cat:
            confirm = await ai_confirm_match(desc_value, po_value, matched_rows, group_cat)
            if confirm.get("match", False):
                results.append({
                    "assigned_category": group_cat,
                    "method": "group_match_confirmed",
                    "reasoning": confirm.get("reasoning", ""),
                    "matched_rows": matched_rows[:3],
                })
                continue
        ai_result = await ai_classify_new(desc_value, po_value)
        results.append({
            "assigned_category": ai_result.get("category", "UNKNOWN"),
            "method": "ai_new_classification",
            "reasoning": ai_result.get("reasoning", ""),
            "matched_rows": matched_rows[:3] if matched_rows else [],
        })
    return results


# ══════════════════════════════════════════════════════════════════════════════
#  DIFF-BASED CHANGE DETECTION & DOWNLOAD
# ══════════════════════════════════════════════════════════════════════════════

def compute_changed_row_indices(df: pd.DataFrame, class_col: str, new_col: str = "New Classification"):
    """Rows where New Classification is non-empty (i.e. was edited)."""
    changed = []
    if new_col not in df.columns:
        return changed
    for i in range(len(df)):
        if str(df.iloc[i][new_col]).strip():
            changed.append(i)
    return changed


def build_download_bytes(df: pd.DataFrame, class_col: str, new_col: str = "New Classification") -> bytes:
    """Export Excel with yellow highlight on rows where old != new classification."""
    changed_indices = compute_changed_row_indices(df, class_col, new_col)
    base = BytesIO()
    df.to_excel(base, index=False, engine="openpyxl")
    base.seek(0)
    wb = load_workbook(base)
    ws = wb.active
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for idx in changed_indices:
        row = idx + 2
        if row <= ws.max_row:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
